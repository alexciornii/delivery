from django.core.management.base import BaseCommand, CommandError
from prj.settings import DATA_DIR
from openpyxl import load_workbook
from apps.market.models import Category
from apps.market.models import Product


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB...')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print(f'Start importing from excel {DATA_DIR}')

        wb = load_workbook(DATA_DIR + '/price.xlsx')
        # print(wb.get_sheet_names())
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        print(sheet.max_row)
        for cnt in range(13, 14):
            item = sheet.cell(row=cnt, column=5).value
            group = sheet.cell(row=cnt, column=6).value
            image = sheet.cell(row=cnt, column=2).value
            category = Category.objects.filter(name=group)
            print(item)
            print(group)
            print(image)
            if len(category) == 0:
                category = Category()
                category.name = group
                category.save()
            else:
                category = category.get()

            product = Product()
            product.name = item
            product.category = category
            product.image = image
            product.save()
